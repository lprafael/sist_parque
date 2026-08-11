"""Lee las planillas oficiales PLANILLA DE BAJA DE BUSES {año}.xlsx.

Fuente embebida: app/data/bajas_oficiales.json (viaja con el backend).
Si existe el Excel en BAJA DE BUSES/, se usa ese y pisa el JSON.
"""
from __future__ import annotations

import json
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

_CACHE: dict[tuple[str, float], list[dict[str, Any]]] = {}
_JSON_CACHE: Optional[dict[str, list[dict[str, Any]]]] = None
_JSON_PATH = Path(__file__).resolve().parents[1] / "data" / "bajas_oficiales.json"


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(v).strip().upper())


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _as_date(v: Any) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _json_embebido() -> dict[str, list[dict[str, Any]]]:
    global _JSON_CACHE
    if _JSON_CACHE is not None:
        return _JSON_CACHE
    if not _JSON_PATH.is_file():
        _JSON_CACHE = {}
        return _JSON_CACHE
    raw = json.loads(_JSON_PATH.read_text(encoding="utf-8"))
    out: dict[str, list[dict[str, Any]]] = {}
    for key, rows in (raw or {}).items():
        parsed = []
        for item in rows:
            rec = dict(item)
            fec = rec.get("fecha_baja")
            if isinstance(fec, str) and fec:
                try:
                    rec["fecha_baja"] = date.fromisoformat(fec[:10])
                except ValueError:
                    rec["fecha_baja"] = None
            parsed.append(rec)
        out[str(key)] = parsed
    _JSON_CACHE = out
    return _JSON_CACHE


def dir_planillas() -> Path:
    env = (os.getenv("BAJAS_PLANILLAS_DIR") or "").strip()
    if env and Path(env).is_dir():
        return Path(env)
    here = Path(__file__).resolve()
    cwd = Path.cwd()
    for cand in (
        here.parents[3] / "BAJA DE BUSES",
        here.parents[2] / "BAJA DE BUSES",
        here.parents[1] / "BAJA DE BUSES",
        cwd / "BAJA DE BUSES",
        cwd.parent / "BAJA DE BUSES",
        Path("/app/BAJA DE BUSES"),
    ):
        try:
            if cand.is_dir():
                return cand
        except (OSError, IndexError):
            continue
    return here.parents[2] / "BAJA DE BUSES"


def path_planilla(anio: int) -> Optional[Path]:
    folder = dir_planillas()
    name = f"PLANILLA DE BAJA DE BUSES {anio}.xlsx"
    path = folder / name
    if path.is_file() and not path.name.startswith("~$"):
        return path
    return None


def anios_con_planilla() -> list[int]:
    found: set[int] = set()
    for key in _json_embebido():
        try:
            found.add(int(key))
        except ValueError:
            continue
    folder = dir_planillas()
    if folder.is_dir():
        for p in folder.glob("PLANILLA DE BAJA DE BUSES *.xlsx"):
            if p.name.startswith("~$"):
                continue
            m = re.search(r"(20\d{2})", p.stem)
            if m:
                found.add(int(m.group(1)))
    return sorted(found)


def _col(header, *needles: str) -> Optional[int]:
    for j, h in enumerate(header or []):
        if h is None:
            continue
        key = str(h).strip().upper()
        if any(n in key for n in needles):
            return j
    return None


def leer_planilla(anio: int) -> list[dict[str, Any]]:
    path = path_planilla(anio)
    if not path:
        return list(_json_embebido().get(str(anio), []))
    mtime = path.stat().st_mtime
    cache_key = (str(path), mtime)
    if cache_key in _CACHE:
        return _CACHE[cache_key]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_i, header = None, None
    for i, row in enumerate(rows[:25]):
        joined = " | ".join(str(c).upper() if c is not None else "" for c in row)
        if "CHASSIS" in joined or "CHASIS" in joined:
            header_i, header = i, row
            break
    if header_i is None:
        return []

    i_ord = _col(header, "ORDEN")
    i_marca = _col(header, "MARCA")
    i_anio = _col(header, "AÑO", "ANO")
    i_ch = _col(header, "CHASSIS", "CHASIS")
    i_rua = _col(header, "RUA")
    i_emp = _col(header, "EMPRESA")
    i_fec = _col(header, "FECHA")
    i_mot = _col(header, "MOTIVO")
    i_meu = _col(header, "MEU", "OBSERV")

    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[header_i + 1 :]:
        if all(c is None or str(c).strip() == "" for c in row[:9]):
            continue
        ch_raw = _cell(row[i_ch]) if i_ch is not None else ""
        ch = _norm(ch_raw)
        if not ch or ch in seen:
            continue
        seen.add(ch)
        fec = _as_date(row[i_fec]) if i_fec is not None else None
        items.append(
            {
                "orden": row[i_ord] if i_ord is not None else None,
                "marca": _cell(row[i_marca]) if i_marca is not None else "",
                "anio_bus": row[i_anio] if i_anio is not None else None,
                "chasis": ch_raw,
                "chasis_norm": ch,
                "rua": _cell(row[i_rua]).upper() if i_rua is not None else "",
                "empresa": _cell(row[i_emp]) if i_emp is not None else "",
                "fecha_baja": fec,
                "motivo": _cell(row[i_mot]) if i_mot is not None else "",
                "meu": _cell(row[i_meu]) if i_meu is not None else "",
                "anio_periodo": anio,
            }
        )

    # evict old cache entries for this path
    for k in list(_CACHE):
        if k[0] == str(path) and k != cache_key:
            _CACHE.pop(k, None)
    _CACHE[cache_key] = items
    return items
