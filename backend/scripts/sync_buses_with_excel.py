import openpyxl
import psycopg2
import os

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

excel_ruas = set()
excel_chassis = set()

for r in range(7, 1875):
    rua = ws.cell(row=r, column=5).value
    chassis = ws.cell(row=r, column=4).value
    if rua: excel_ruas.add(str(rua).strip().upper())
    if chassis: excel_chassis.add(str(chassis).strip().upper())

conn = psycopg2.connect(
    host=os.getenv("DB_HOST", "168.90.177.232"),
    port=int(os.getenv("DB_PORT", 2024)),
    user=os.getenv("DB_USER", "cid_admin_user"),
    password=os.getenv("DB_PASSWORD", "vmtdmtcidccm"),
    dbname=os.getenv("DB_NAME", "bbdd-monitoreo-cid"),
    sslmode="disable"
)
cur = conn.cursor()

# 1. Marcar como INACTIVO los buses que no pertenecen a la planilla oficial 2026
cur.execute("SELECT id_bus, rua, numero_chassis FROM registro_habilitacion.buses;")
db_buses = cur.fetchall()

inactivos_cnt = 0
activos_cnt = 0

for id_b, rua_b, ch_b in db_buses:
    rua_s = str(rua_b).strip().upper() if rua_b else ""
    ch_s = str(ch_b).strip().upper() if ch_b else ""
    
    if rua_s in excel_ruas or ch_s in excel_chassis:
        cur.execute("UPDATE registro_habilitacion.buses SET estado_bus = 'ACTIVO' WHERE id_bus = %s;", (id_b,))
        activos_cnt += 1
    else:
        cur.execute("UPDATE registro_habilitacion.buses SET estado_bus = 'INACTIVO' WHERE id_bus = %s;", (id_b,))
        inactivos_cnt += 1

conn.commit()

print("--- SINCRONIZACIÓN DE ESTADO DE BUSES COMPLETADA ---")
print(f"Buses Activos (en Planilla Oficial 2026) : {activos_cnt}")
print(f"Buses Inactivos / Obsoletos fuera de Excel: {inactivos_cnt}")

cur.close()
conn.close()
