"""Parser y sincronización de la planilla ITV (hoja General)."""
from __future__ import annotations

import asyncio
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Bus,
    ItvBus,
    Marca,
    MarcaCarroceria,
    SeguroBus,
    TipoCarroceria,
    TipoSeguro,
    TipoServicio,
)

SCHEMA = "registro_habilitacion"
DATA_START = 7
HEADER_SCAN_MAX = 12

# Alias de encabezados → clave interna (normalizados sin tildes/espacios)
HEADER_ALIASES: dict[str, str] = {
    "tg": "tg",
    "fecharegistro": "fecha_registro",
    "ndeorden": "numero_orden",
    "nodeorden": "numero_orden",
    "nrodeorden": "numero_orden",
    "marca": "marca",
    "ano": "anio",
    "año": "anio",
    "nchassis": "chassis",
    "nochassis": "chassis",
    "nrochassis": "chassis",
    "rua": "rua",
    "podrtd": "pod_rtd",
    "documentos": "documentos",
    "habilitacion": "habilitacion",
    "seguropasajeros": "seguro_pasajeros",
    "seguroterceros": "seguro_terceros",
    "tipodeservicio": "tipo_servicio",
    "situaciondebus": "situacion_bus",
    "fechabase": "fecha_base",
    "tegnologiadebus": "tecnologia",
    "tegnoologiadebus": "tecnologia",
    "tecnologiadebus": "tecnologia",
    "tipodecarroceria": "tipo_carroceria",
    "tipobus": "tipo_bus",
    "marcadecarroceria": "marca_carroceria",
    "fechadeivencimientodelitvanterior": "itv_anterior",
    "fechadevencimientodelitvanterior": "itv_anterior",
    "fechadeitv": "fecha_itv",
    "vencimientodeitv": "vencimiento_itv",
    "situaciondeitvaprobada": "resultado_itv",
    "situaciondeitvaprobadaovencida": "situacion_itv",
    "zonal": "zonal",
    "codigo": "codigo",
    "empresalinea": "empresa_linea",
    "taller": "taller",
}


def _norm_header(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def parse_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if val in ("", "00/00/00", "N/A", "N/D", "-", "None", "null"):
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val[:19], fmt).date()
            except ValueError:
                continue
    return None


def clean_str(val: Any, upper: bool = False) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "N/D", "N/A", "-", "None", "null"):
        return None
    return s.upper() if upper else s


def normalize_tipo_servicio(val: Any) -> Optional[str]:
    raw = clean_str(val, upper=True)
    if not raw:
        return None
    s = raw.replace("(*)", "").replace("*", "").strip()
    s = re.sub(r"\s+", " ", s)
    if "ELECTR" in s:
        return "ELECTRICO"
    if "DIFER" in s:
        return "DIFERENCIADO"
    if "CONVEN" in s:
        return "CONVENCIONAL"
    return s[:100]


def normalize_resultado_itv(val: Any) -> Optional[str]:
    raw = clean_str(val, upper=True)
    if not raw:
        return None
    if "PARCIAL" in raw:
        return "PARCIAL"
    if "TOTAL" in raw:
        return "TOTAL"
    return None


@dataclass
class ExcelBusRow:
    row_num: int
    numero_orden: Optional[int] = None
    marca: Optional[str] = None
    anio: Optional[int] = None
    chassis: Optional[str] = None
    rua: Optional[str] = None
    pod_rtd: Optional[str] = None
    documentos: Optional[str] = None
    habilitacion: Optional[date] = None
    seguro_pasajeros: Optional[date] = None
    seguro_terceros: Optional[date] = None
    tipo_servicio: Optional[str] = None
    tecnologia: Optional[str] = None
    tipo_carroceria: Optional[str] = None
    marca_carroceria: Optional[str] = None
    itv_anterior: Optional[date] = None
    fecha_itv: Optional[date] = None
    vencimiento_itv: Optional[date] = None
    resultado_itv: Optional[str] = None
    situacion_itv: Optional[str] = None
    codigo: Optional[int] = None
    empresa_linea: Optional[str] = None
    taller: Optional[str] = None


@dataclass
class PreviewResult:
    hoja: str
    fecha_corte: Optional[str]
    total_excel: int
    matched_rua: int = 0
    matched_chassis: int = 0
    solo_excel: int = 0
    solo_db_activos: int = 0
    itv_actualizar: int = 0
    itv_igual: int = 0
    itv_sin_fecha: int = 0
    con_seguro_pasajeros: int = 0
    con_seguro_terceros: int = 0
    tipos_servicio: dict[str, int] = field(default_factory=dict)
    muestra_solo_excel: list[dict] = field(default_factory=list)
    muestra_solo_db: list[dict] = field(default_factory=list)
    muestra_itv_diff: list[dict] = field(default_factory=list)
    errores_parseo: list[str] = field(default_factory=list)


@dataclass
class ApplyResult:
    buses_creados: int = 0
    buses_actualizados: int = 0
    buses_activados: int = 0
    buses_inactivados: int = 0
    itv_insertados: int = 0
    itv_sin_cambio: int = 0
    seguros_insertados: int = 0
    auxiliar_filas: int = 0
    errores: list[str] = field(default_factory=list)


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Devuelve (fila_header_1based, mapa clave→índice_columna_0based)."""
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_MAX, values_only=True), start=1):
        mapping: dict[str, int] = {}
        for c_idx, cell in enumerate(row):
            key = HEADER_ALIASES.get(_norm_header(cell))
            if key and key not in mapping:
                mapping[key] = c_idx
        if "rua" in mapping and "chassis" in mapping and ("marca" in mapping or "numero_orden" in mapping):
            return r_idx, mapping
    raise ValueError(
        "No se encontró la fila de encabezados de la planilla ITV "
        "(se esperan columnas RUA, Nº CHASSIS, MARCA / Nº DE ORDEN)."
    )


def parse_general_sheet(file_bytes: bytes) -> tuple[str, Optional[str], list[ExcelBusRow], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet_name = "General" if "General" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    header_row, colmap = _find_header_row(ws)
    errors: list[str] = []
    rows: list[ExcelBusRow] = []
    fecha_corte: Optional[str] = None

    def cell(row_tuple, key: str):
        idx = colmap.get(key)
        if idx is None or idx >= len(row_tuple):
            return None
        return row_tuple[idx]

    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue

        rua = clean_str(cell(row, "rua"), upper=True)
        chassis = clean_str(cell(row, "chassis"), upper=True)
        if not rua and not chassis:
            continue
        if rua and rua.startswith("#"):
            continue

        anio_raw = cell(row, "anio")
        anio = None
        if isinstance(anio_raw, int):
            anio = anio_raw
        elif anio_raw is not None:
            try:
                anio = int(str(anio_raw).strip()[:4])
            except ValueError:
                pass

        orden_raw = cell(row, "numero_orden")
        orden = orden_raw if isinstance(orden_raw, int) else None
        if orden is None and orden_raw is not None:
            try:
                orden = int(float(str(orden_raw).strip()))
            except ValueError:
                pass

        codigo_raw = cell(row, "codigo")
        codigo = codigo_raw if isinstance(codigo_raw, int) else None
        if codigo is None and codigo_raw is not None:
            try:
                codigo = int(float(str(codigo_raw).strip()))
            except ValueError:
                pass

        fb = parse_date(cell(row, "fecha_base"))
        if fb and not fecha_corte:
            fecha_corte = fb.isoformat()

        item = ExcelBusRow(
            row_num=r_idx,
            numero_orden=orden,
            marca=clean_str(cell(row, "marca")),
            anio=anio,
            chassis=chassis,
            rua=rua,
            pod_rtd=clean_str(cell(row, "pod_rtd"), upper=True),
            documentos=clean_str(cell(row, "documentos")),
            habilitacion=parse_date(cell(row, "habilitacion")),
            seguro_pasajeros=parse_date(cell(row, "seguro_pasajeros")),
            seguro_terceros=parse_date(cell(row, "seguro_terceros")),
            tipo_servicio=normalize_tipo_servicio(cell(row, "tipo_servicio")),
            tecnologia=clean_str(cell(row, "tecnologia"), upper=True),
            tipo_carroceria=clean_str(cell(row, "tipo_carroceria")),
            marca_carroceria=clean_str(cell(row, "marca_carroceria")),
            itv_anterior=parse_date(cell(row, "itv_anterior")),
            fecha_itv=parse_date(cell(row, "fecha_itv")),
            vencimiento_itv=parse_date(cell(row, "vencimiento_itv")),
            resultado_itv=normalize_resultado_itv(cell(row, "resultado_itv")),
            situacion_itv=clean_str(cell(row, "situacion_itv"), upper=True),
            codigo=codigo,
            empresa_linea=clean_str(cell(row, "empresa_linea")),
            taller=clean_str(cell(row, "taller")),
        )
        rows.append(item)

    wb.close()
    return sheet_name, fecha_corte, rows, errors


async def build_preview(db: AsyncSession, file_bytes: bytes) -> PreviewResult:
    hoja, fecha_corte, rows, errors = await asyncio.to_thread(parse_general_sheet, file_bytes)
    preview = PreviewResult(
        hoja=hoja,
        fecha_corte=fecha_corte,
        total_excel=len(rows),
        errores_parseo=errors,
    )

    for r in rows:
        key = r.tipo_servicio or "(vacío)"
        preview.tipos_servicio[key] = preview.tipos_servicio.get(key, 0) + 1
        if r.seguro_pasajeros:
            preview.con_seguro_pasajeros += 1
        if r.seguro_terceros:
            preview.con_seguro_terceros += 1
        if not r.vencimiento_itv:
            preview.itv_sin_fecha += 1

    result = await db.execute(select(Bus.id_bus, Bus.rua, Bus.numero_chassis, Bus.estado_bus))
    db_buses = result.all()

    rua_map: dict[str, Any] = {}
    chassis_map: dict[str, Any] = {}
    activos_ids: set[int] = set()
    for id_bus, rua, chassis, estado in db_buses:
        if rua:
            rua_map[str(rua).strip().upper()] = id_bus
        if chassis:
            chassis_map[str(chassis).strip().upper()] = id_bus
        if (estado or "").upper() == "ACTIVO":
            activos_ids.add(id_bus)

    itv_res = await db.execute(
        select(ItvBus.id_bus, ItvBus.fecha_itv, ItvBus.fecha_vencimiento).where(ItvBus.es_vigente.is_(True))
    )
    itv_map = {row[0]: (row[1], row[2]) for row in itv_res.all()}

    excel_ids: set[int] = set()
    matched_ids: set[int] = set()

    for r in rows:
        id_bus = None
        how = None
        if r.rua and r.rua in rua_map:
            id_bus = rua_map[r.rua]
            how = "rua"
            preview.matched_rua += 1
        elif r.chassis and r.chassis in chassis_map:
            id_bus = chassis_map[r.chassis]
            how = "chassis"
            preview.matched_chassis += 1

        if id_bus:
            matched_ids.add(id_bus)
            excel_ids.add(id_bus)
            cur_itv = itv_map.get(id_bus)
            if r.vencimiento_itv:
                if not cur_itv or cur_itv[0] != r.fecha_itv or cur_itv[1] != r.vencimiento_itv:
                    preview.itv_actualizar += 1
                    if len(preview.muestra_itv_diff) < 15:
                        preview.muestra_itv_diff.append({
                            "rua": r.rua,
                            "chassis": r.chassis,
                            "itv_db": cur_itv[1].isoformat() if cur_itv and cur_itv[1] else None,
                            "itv_excel": r.vencimiento_itv.isoformat(),
                            "match": how,
                        })
                else:
                    preview.itv_igual += 1
        else:
            preview.solo_excel += 1
            if len(preview.muestra_solo_excel) < 15:
                preview.muestra_solo_excel.append({
                    "fila": r.row_num,
                    "rua": r.rua,
                    "chassis": r.chassis,
                    "marca": r.marca,
                    "anio": r.anio,
                    "empresa": r.empresa_linea,
                })

    solo_db = activos_ids - matched_ids
    preview.solo_db_activos = len(solo_db)

    id_to_meta = {
        id_bus: (rua, chassis)
        for id_bus, rua, chassis, _ in db_buses
    }
    for id_bus in list(solo_db)[:15]:
        rua, chassis = id_to_meta.get(id_bus, (None, None))
        preview.muestra_solo_db.append({"id_bus": id_bus, "rua": rua, "chassis": chassis})

    return preview


def func_upper_eq(column, value: str):
    from sqlalchemy import func
    return func.upper(func.trim(column)) == value.strip().upper()


async def _scalars_first(db: AsyncSession, stmt):
    res = await db.execute(stmt)
    return res.scalars().first()


async def _get_or_create_marca(db: AsyncSession, cache: dict, nombre: Optional[str]) -> Optional[int]:
    if not nombre:
        return None
    key = nombre.strip().upper()
    if key in cache:
        return cache[key]
    m = await _scalars_first(db, select(Marca).where(func_upper_eq(Marca.nombre, nombre)))
    if m:
        cache[key] = m.id_marca
        return m.id_marca
    m = Marca(nombre=nombre.strip()[:100])
    db.add(m)
    await db.flush()
    cache[key] = m.id_marca
    return m.id_marca


async def _get_or_create_marca_carr(db: AsyncSession, cache: dict, nombre: Optional[str]) -> Optional[int]:
    if not nombre:
        return None
    key = nombre.strip().upper()
    if key in cache:
        return cache[key]
    m = await _scalars_first(
        db, select(MarcaCarroceria).where(func_upper_eq(MarcaCarroceria.nombre, nombre))
    )
    if m:
        cache[key] = m.id_marca_carroceria
        return m.id_marca_carroceria
    m = MarcaCarroceria(nombre=nombre.strip()[:100])
    db.add(m)
    await db.flush()
    cache[key] = m.id_marca_carroceria
    return m.id_marca_carroceria


async def _get_or_create_tipo_carr(db: AsyncSession, cache: dict, desc: Optional[str]) -> Optional[int]:
    if not desc:
        return None
    key = str(desc).strip().upper()
    if key in cache:
        return cache[key]
    d = str(desc).strip()[:100]
    m = await _scalars_first(
        db, select(TipoCarroceria).where(func_upper_eq(TipoCarroceria.descripcion, d))
    )
    if m:
        cache[key] = m.id_tipo
        return m.id_tipo
    m = TipoCarroceria(descripcion=d)
    db.add(m)
    await db.flush()
    cache[key] = m.id_tipo
    return m.id_tipo


async def _resolve_tipo_servicio(db: AsyncSession, cache: dict, nombre: Optional[str]) -> Optional[int]:
    if not nombre:
        return None
    if nombre in cache:
        return cache[nombre]
    m = await _scalars_first(db, select(TipoServicio).where(TipoServicio.nombre == nombre))
    if m:
        cache[nombre] = m.id_tipo_servicio
        return m.id_tipo_servicio
    m = TipoServicio(nombre=nombre, descripcion=nombre, activo=True)
    db.add(m)
    await db.flush()
    cache[nombre] = m.id_tipo_servicio
    return m.id_tipo_servicio


async def _tipo_seguro_ids(db: AsyncSession) -> dict[str, int]:
    res = await db.execute(select(TipoSeguro))
    out = {}
    for t in res.scalars().all():
        out[t.nombre.upper()] = t.id_tipo_seguro
    for required in ("PASAJEROS", "TERCEROS"):
        if required not in out:
            t = TipoSeguro(nombre=required, descripcion=required, activo=True)
            db.add(t)
            await db.flush()
            out[required] = t.id_tipo_seguro
    return out


async def _list_itv_vigentes(db: AsyncSession, id_bus: int) -> list[ItvBus]:
    res = await db.execute(
        select(ItvBus)
        .where(ItvBus.id_bus == id_bus, ItvBus.es_vigente.is_(True))
        .order_by(ItvBus.fecha_vencimiento.desc(), ItvBus.id_itv.desc())
    )
    return list(res.scalars().all())


async def _list_seguros_vigentes(db: AsyncSession, id_bus: int, id_tipo: int) -> list[SeguroBus]:
    res = await db.execute(
        select(SeguroBus)
        .where(
            SeguroBus.id_bus == id_bus,
            SeguroBus.id_tipo_seguro == id_tipo,
            SeguroBus.seguro_vigente.is_(True),
        )
        .order_by(SeguroBus.fecha_vencimiento.desc(), SeguroBus.id_seguro.desc())
    )
    return list(res.scalars().all())


def _bus_in_excel(bus: Bus, excel_ruas: set[str], excel_chassis: set[str]) -> bool:
    rua = str(bus.rua).strip().upper() if bus.rua else ""
    chassis = str(bus.numero_chassis).strip().upper() if bus.numero_chassis else ""
    return (rua and rua in excel_ruas) or (chassis and chassis in excel_chassis)


async def sincronizar_estado_desde_excel(
    db: AsyncSession,
    file_bytes: bytes,
    *,
    inactivar_fuera: bool = True,
) -> ApplyResult:
    """Solo alinea ACTIVO/INACTIVO según RUA/chasis del Excel (recuperación segura)."""
    _, _, rows, _ = await asyncio.to_thread(parse_general_sheet, file_bytes)
    result = ApplyResult()
    excel_ruas = {r.rua for r in rows if r.rua}
    excel_chassis = {r.chassis for r in rows if r.chassis}

    bus_res = await db.execute(select(Bus))
    buses = list(bus_res.scalars().all())
    for b in buses:
        if _bus_in_excel(b, excel_ruas, excel_chassis):
            if (b.estado_bus or "").upper() != "ACTIVO":
                b.estado_bus = "ACTIVO"
                result.buses_activados += 1
        elif inactivar_fuera and (b.estado_bus or "").upper() == "ACTIVO":
            b.estado_bus = "INACTIVO"
            result.buses_inactivados += 1

    # commit lo hace get_db
    return result


async def apply_import(
    db: AsyncSession,
    file_bytes: bytes,
    *,
    sincronizar_estado: bool = True,
    crear_faltantes: bool = True,
    usuario: Optional[str] = None,
) -> ApplyResult:
    hoja, fecha_corte, rows, _ = await asyncio.to_thread(parse_general_sheet, file_bytes)
    result = ApplyResult()
    today = date.today()

    excel_ruas = {r.rua for r in rows if r.rua}
    excel_chassis = {r.chassis for r in rows if r.chassis}

    bus_res = await db.execute(select(Bus))
    buses = list(bus_res.scalars().all())
    rua_map: dict[str, Bus] = {}
    chassis_map: dict[str, Bus] = {}
    for b in buses:
        if b.rua:
            rua_map[str(b.rua).strip().upper()] = b
        if b.numero_chassis:
            chassis_map[str(b.numero_chassis).strip().upper()] = b

    # Precarga ITV / seguros vigentes para evitar N+1 y scalar_one
    itv_res = await db.execute(
        select(ItvBus)
        .where(ItvBus.es_vigente.is_(True))
        .order_by(ItvBus.fecha_vencimiento.desc(), ItvBus.id_itv.desc())
    )
    itv_by_bus: dict[int, list[ItvBus]] = {}
    for itv in itv_res.scalars().all():
        itv_by_bus.setdefault(itv.id_bus, []).append(itv)

    seg_res = await db.execute(
        select(SeguroBus)
        .where(SeguroBus.seguro_vigente.is_(True))
        .order_by(SeguroBus.fecha_vencimiento.desc(), SeguroBus.id_seguro.desc())
    )
    seg_by_bus_tipo: dict[tuple[int, int], list[SeguroBus]] = {}
    for seg in seg_res.scalars().all():
        seg_by_bus_tipo.setdefault((seg.id_bus, seg.id_tipo_seguro), []).append(seg)

    marca_cache: dict = {}
    marca_carr_cache: dict = {}
    tipo_carr_cache: dict = {}
    tipo_serv_cache: dict = {}
    tipo_seguro = await _tipo_seguro_ids(db)

    seen_ids: set[int] = set()
    auxiliar_rows: list[dict] = []
    max_errores_guardados = 80

    for r in rows:
        try:
            async with db.begin_nested():
                id_marca = await _get_or_create_marca(db, marca_cache, r.marca)
                id_marca_carr = await _get_or_create_marca_carr(db, marca_carr_cache, r.marca_carroceria)
                id_tipo_carr = await _get_or_create_tipo_carr(db, tipo_carr_cache, r.tipo_carroceria)
                id_tipo_serv = await _resolve_tipo_servicio(db, tipo_serv_cache, r.tipo_servicio)

                bus: Optional[Bus] = None
                if r.rua and r.rua in rua_map:
                    bus = rua_map[r.rua]
                elif r.chassis and r.chassis in chassis_map:
                    bus = chassis_map[r.chassis]

                if bus:
                    bus.numero_orden = r.numero_orden if r.numero_orden is not None else bus.numero_orden
                    if id_marca:
                        bus.id_marca = id_marca
                    if r.anio:
                        bus.año = r.anio
                    if id_tipo_carr:
                        bus.id_tipo_carroceria = id_tipo_carr
                    if id_marca_carr:
                        bus.id_marca_carroceria = id_marca_carr
                    if id_tipo_serv:
                        bus.id_tipo_servicio = id_tipo_serv
                    if r.tecnologia:
                        bus.combustible = r.tecnologia[:50]
                    if r.rua and (not bus.rua or str(bus.rua).upper() in ("N/D",)):
                        bus.rua = r.rua
                    result.buses_actualizados += 1
                elif crear_faltantes:
                    rua_val = (r.rua or f"CHASSIS_{r.chassis}")[:20]
                    chassis_val = (r.chassis or f"RUA_{r.rua}")[:50]
                    bus = Bus(
                        numero_orden=r.numero_orden,
                        id_marca=id_marca,
                        año=r.anio or 2000,
                        numero_chassis=chassis_val,
                        rua=rua_val,
                        id_tipo_carroceria=id_tipo_carr,
                        id_marca_carroceria=id_marca_carr,
                        id_tipo_servicio=id_tipo_serv,
                        combustible=(r.tecnologia[:50] if r.tecnologia else None),
                        estado_bus="ACTIVO",
                    )
                    db.add(bus)
                    await db.flush()
                    rua_map[rua_val.upper()] = bus
                    chassis_map[chassis_val.upper()] = bus
                    result.buses_creados += 1
                else:
                    continue

                seen_ids.add(bus.id_bus)

                if r.vencimiento_itv:
                    vigentes = itv_by_bus.get(bus.id_bus, [])
                    current = vigentes[0] if vigentes else None
                    fecha_itv = r.fecha_itv or r.vencimiento_itv
                    same = (
                        current is not None
                        and current.fecha_itv == fecha_itv
                        and current.fecha_vencimiento == r.vencimiento_itv
                        and (current.resultado_itv or None) == (r.resultado_itv or current.resultado_itv)
                        and (current.centro_itv or None) == (r.taller or current.centro_itv)
                    )
                    if same:
                        for extra in vigentes[1:]:
                            extra.es_vigente = False
                        itv_by_bus[bus.id_bus] = [current]
                        result.itv_sin_cambio += 1
                    else:
                        for itv in vigentes:
                            itv.es_vigente = False
                        nuevo = ItvBus(
                            id_bus=bus.id_bus,
                            fecha_itv=fecha_itv,
                            fecha_vencimiento=r.vencimiento_itv,
                            resultado_itv=r.resultado_itv,
                            centro_itv=(r.taller[:200] if r.taller else None),
                            es_vigente=True,
                        )
                        db.add(nuevo)
                        itv_by_bus[bus.id_bus] = [nuevo]
                        result.itv_insertados += 1

                for nombre_tipo, fecha_venc in (
                    ("PASAJEROS", r.seguro_pasajeros),
                    ("TERCEROS", r.seguro_terceros),
                ):
                    if not fecha_venc:
                        continue
                    id_tipo = tipo_seguro[nombre_tipo]
                    key = (bus.id_bus, id_tipo)
                    vigentes_seg = seg_by_bus_tipo.get(key, [])
                    cur_seg = vigentes_seg[0] if vigentes_seg else None
                    if cur_seg and cur_seg.fecha_vencimiento == fecha_venc:
                        for extra in vigentes_seg[1:]:
                            extra.seguro_vigente = False
                        seg_by_bus_tipo[key] = [cur_seg]
                        continue
                    for seg in vigentes_seg:
                        seg.seguro_vigente = False
                    nuevo_seg = SeguroBus(
                        id_bus=bus.id_bus,
                        id_tipo_seguro=id_tipo,
                        fecha_inicio=today,
                        fecha_vencimiento=fecha_venc,
                        seguro_vigente=True,
                    )
                    db.add(nuevo_seg)
                    seg_by_bus_tipo[key] = [nuevo_seg]
                    result.seguros_insertados += 1

                auxiliar_rows.append({
                    "orden": r.numero_orden,
                    "marca": r.marca,
                    "año": r.anio,
                    "chasis": r.chassis,
                    "RUA": r.rua or (r.chassis or f"ROW{r.row_num}"),
                    "POD / RTD": r.pod_rtd,
                    "Documentos": r.documentos,
                    "Habilitacion": r.habilitacion,
                    "Seguro Pasajeros": r.seguro_pasajeros,
                    "Seguro Terceros": r.seguro_terceros,
                    "Tipo de Servicio": r.tipo_servicio,
                    "Tipo de Carroceria": r.tipo_carroceria,
                    "Marca de Carroceria": r.marca_carroceria,
                    "Fecha de Vencimiento del ITV Anterior": r.itv_anterior,
                    "Fecha de ITV": r.fecha_itv,
                    "Vencimiento de ITV": r.vencimiento_itv,
                    "Situacion de ITV Aprobada": r.resultado_itv or r.situacion_itv,
                    "Codigo": r.codigo,
                    "Empresa_Linea": r.empresa_linea,
                    "id_marca": id_marca,
                    "id_marca_carroceria": id_marca_carr,
                    "id_tipo_carroceria": id_tipo_carr,
                })

        except Exception as exc:  # noqa: BLE001
            if len(result.errores) < max_errores_guardados:
                result.errores.append(f"Fila {r.row_num} RUA={r.rua}: {exc}")
            elif len(result.errores) == max_errores_guardados:
                result.errores.append("… (más errores omitidos en la respuesta)")
            continue

    # Releer buses por si hubo creates
    bus_res = await db.execute(select(Bus))
    buses = list(bus_res.scalars().all())

    # ACTIVO/INACTIVO según presencia en Excel (no según filas OK)
    if sincronizar_estado:
        for b in buses:
            if _bus_in_excel(b, excel_ruas, excel_chassis):
                if (b.estado_bus or "").upper() != "ACTIVO":
                    b.estado_bus = "ACTIVO"
                    result.buses_activados += 1
            elif (b.estado_bus or "").upper() == "ACTIVO":
                b.estado_bus = "INACTIVO"
                result.buses_inactivados += 1

    # Refrescar staging auxiliar (best-effort)
    try:
        await db.execute(text(f"TRUNCATE TABLE {SCHEMA}.auxiliar"))
        for ar in auxiliar_rows:
            await db.execute(
                text(
                    f"""
                    INSERT INTO {SCHEMA}.auxiliar (
                        orden, marca, año, chasis, "RUA",
                        "POD / RTD", "Documentos", "Habilitacion",
                        "Seguro Pasajeros", "Seguro Terceros",
                        "Tipo de Servicio", "Tipo de Carroceria", "Marca de Carroceria",
                        "Fecha de Vencimiento del ITV Anterior", "Fecha de ITV",
                        "Vencimiento de ITV", "Situacion de ITV Aprobada",
                        "Codigo", "Empresa_Linea",
                        id_marca, id_marca_carroceria, id_tipo_carroceria
                    ) VALUES (
                        :orden, :marca, :anio, :chasis, :rua,
                        :pod, :docs, :hab,
                        :seg_pas, :seg_ter,
                        :tipo_serv, :tipo_carr, :marca_carr,
                        :itv_ant, :fecha_itv,
                        :venc_itv, :sit_itv,
                        :codigo, :empresa,
                        :id_marca, :id_marca_carr, :id_tipo_carr
                    )
                    """
                ),
                {
                    "orden": ar["orden"],
                    "marca": ar["marca"],
                    "anio": ar["año"],
                    "chasis": ar["chasis"],
                    "rua": ar["RUA"],
                    "pod": ar["POD / RTD"],
                    "docs": ar["Documentos"],
                    "hab": ar["Habilitacion"],
                    "seg_pas": ar["Seguro Pasajeros"],
                    "seg_ter": ar["Seguro Terceros"],
                    "tipo_serv": ar["Tipo de Servicio"],
                    "tipo_carr": ar["Tipo de Carroceria"],
                    "marca_carr": ar["Marca de Carroceria"],
                    "itv_ant": ar["Fecha de Vencimiento del ITV Anterior"],
                    "fecha_itv": ar["Fecha de ITV"],
                    "venc_itv": ar["Vencimiento de ITV"],
                    "sit_itv": ar["Situacion de ITV Aprobada"],
                    "codigo": ar["Codigo"],
                    "empresa": ar["Empresa_Linea"],
                    "id_marca": ar["id_marca"],
                    "id_marca_carr": ar["id_marca_carroceria"],
                    "id_tipo_carr": ar["id_tipo_carroceria"],
                },
            )
        result.auxiliar_filas = len(auxiliar_rows)
    except Exception as exc:  # noqa: BLE001
        if len(result.errores) < max_errores_guardados:
            result.errores.append(f"auxiliar (staging): {exc}")

    _ = (hoja, fecha_corte, usuario, seen_ids)
    return result
