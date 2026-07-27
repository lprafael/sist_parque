import openpyxl

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
sheet = wb.active

headers = [sheet.cell(row=6, column=c).value for c in range(1, 20)]

print("Headers (Row 6):")
for i, h in enumerate(headers, 0):
    print(f" Index {i} (Col {i+1}): {h}")

print("\n--- Data Rows 7 to 16 ---")
for r in range(7, 17):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
    rua = row_vals[4]
    itv_ant = row_vals[13]
    fecha_itv = row_vals[14]
    venc_itv = row_vals[15]
    print(f"Row {r} [RUA: {rua}]:")
    print(f"   Col 14 (Index 13 - ITV Anterior): {itv_ant}")
    print(f"   Col 15 (Index 14 - Fecha ITV):    {fecha_itv}")
    print(f"   Col 16 (Index 15 - Venc. ITV):    {venc_itv}")
