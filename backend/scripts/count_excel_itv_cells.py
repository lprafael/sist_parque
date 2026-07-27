import openpyxl
from datetime import datetime, date

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

itv_dates = []
empty_itv = 0

def parse_date(val):
    if val is None: return None
    if isinstance(val, (date, datetime)): return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        val = val.strip()
        if val in ("00/00/00", "", "N/A", "-", "None"): return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
            try: return datetime.strptime(val, fmt).date()
            except ValueError: continue
    return None

for r in range(7, 1875):
    rua = ws.cell(row=r, column=5).value
    chassis = ws.cell(row=r, column=4).value
    venc = ws.cell(row=r, column=16).value # Col 16 = VENCIMIENTO DE ITV
    
    p_venc = parse_date(venc)
    if p_venc:
        itv_dates.append((r, str(rua), p_venc))
    else:
        empty_itv += 1

print(f"Total Buses in Excel (Rows 7..1874): {1875 - 7}")
print(f"Total Buses WITH valid VENCIMIENTO DE ITV date: {len(itv_dates)}")
print(f"Total Buses WITHOUT valid ITV date (or 00/00/00): {empty_itv}")

print("\nSample 10 Parsed ITV Vencimiento Dates:")
for item in itv_dates[:10]:
    print(" ", item)
