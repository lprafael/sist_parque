import io
from datetime import date
from pathlib import Path
from typing import Optional, List, Dict, Any, Callable

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.security import get_current_user, require_roles
from app.models import Bus, BusEmpresa, Eot, TipoServicio
from app.api.v1.endpoints.buses import calcular_estado_itv
from app.services import planilla_viewer

router = APIRouter(prefix="/reportes", tags=["Reportes"])

LOGO_PATH = Path(__file__).resolve().parents[3] / "static" / "Logo-MOPC-VMT.png"

# ── Campos exportables ─────────────────────────────────────
CAMPOS: Dict[str, Dict[str, Any]] = {
    "id_bus":              {"label": "N° ID",              "default": True},
    "numero_orden":        {"label": "N° Orden",           "default": True},
    "rua":                 {"label": "RUA / Placa",        "default": True},
    "numero_chassis":      {"label": "Chasis",             "default": True},
    "año":                 {"label": "Año",                "default": True},
    "antiguedad":          {"label": "Antigüedad (años)",  "default": False},
    "marca":               {"label": "Marca",              "default": True},
    "tipo_carroceria":     {"label": "Tipo Carrocería",    "default": True},
    "marca_carroceria":    {"label": "Marca Carrocería",   "default": True},
    "empresa":             {"label": "Empresa",            "default": True},
    "combustible":         {"label": "Combustible",        "default": True},
    "capacidad_pasajeros": {"label": "Capacidad",          "default": False},
    "cilindrada":          {"label": "Cilindrada",         "default": False},
    "color":               {"label": "Color",              "default": False},
    "tipo_servicio":       {"label": "Tipo Servicio",      "default": False},
    "estado_bus":          {"label": "Estado Bus",         "default": True},
    "tiene_rampa":         {"label": "Rampa discapacitados", "default": False},
    "itv_vencimiento":     {"label": "ITV Vencimiento",    "default": True},
    "itv_estado":          {"label": "Estado ITV",         "default": True},
    "fecha_itv":           {"label": "Fecha ITV",          "default": False},
}

# ── Resúmenes disponibles ──────────────────────────────────
RESUMENES: Dict[str, Dict[str, Any]] = {
    "total":                {"label": "Cantidad total de buses",          "default": True},
    "promedio_antiguedad":  {"label": "Promedio de antigüedad",           "default": True},
    "antiguedad_max":       {"label": "Antigüedad máxima",                "default": True},
    "antiguedad_min":       {"label": "Antigüedad mínima",                "default": False},
    "itv_vencido":          {"label": "Buses con ITV vencido",            "default": True},
    "itv_por_vencer":       {"label": "Buses con ITV por vencer",         "default": True},
    "itv_critico":          {"label": "Buses con ITV crítico",            "default": False},
    "itv_vigente":          {"label": "Buses con ITV vigente",            "default": False},
    "sin_itv":              {"label": "Buses sin ITV",                    "default": False},
    "activos":              {"label": "Buses activos",                    "default": False},
    "inactivos":            {"label": "Buses inactivos",                  "default": False},
}


def _parse_list(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [v.strip() for v in value.split(",") if v.strip()]


def _valor_campo(row: Dict[str, Any], key: str) -> Any:
    val = row.get(key)
    if val is None or val == "":
        return "-"
    if key in ("itv_vencimiento", "fecha_itv") and hasattr(val, "isoformat"):
        return val.isoformat()
    return val


def _calcular_resumenes(rows: List[Dict[str, Any]], keys: List[str]) -> List[tuple]:
    anio_actual = date.today().year
    antiguedades = [anio_actual - r["año"] for r in rows if isinstance(r.get("año"), int)]
    estados = [r.get("itv_estado") for r in rows]

    calculators: Dict[str, Callable[[], Any]] = {
        "total":               lambda: len(rows),
        "promedio_antiguedad": lambda: round(sum(antiguedades) / len(antiguedades), 1) if antiguedades else 0,
        "antiguedad_max":      lambda: max(antiguedades) if antiguedades else 0,
        "antiguedad_min":      lambda: min(antiguedades) if antiguedades else 0,
        "itv_vencido":         lambda: sum(1 for e in estados if e == "VENCIDO"),
        "itv_por_vencer":      lambda: sum(1 for e in estados if e == "POR_VENCER"),
        "itv_critico":         lambda: sum(1 for e in estados if e == "CRITICO"),
        "itv_vigente":         lambda: sum(1 for e in estados if e == "VIGENTE"),
        "sin_itv":             lambda: sum(1 for e in estados if e == "SIN_ITV"),
        "activos":             lambda: sum(1 for r in rows if (r.get("estado_bus") or "").upper() == "ACTIVO"),
        "inactivos":           lambda: sum(1 for r in rows if (r.get("estado_bus") or "").upper() == "INACTIVO"),
    }

    result = []
    for key in keys:
        if key in RESUMENES and key in calculators:
            label = RESUMENES[key]["label"]
            value = calculators[key]()
            if key in ("promedio_antiguedad", "antiguedad_max", "antiguedad_min"):
                value = f"{value} años"
            result.append((label, value))
    return result


def _insertar_logo(ws) -> int:
    """Inserta el logo institucional como encabezado. Retorna la fila donde empieza el contenido."""
    # Filas 1-3 reservadas para logo + título
    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 8
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    title_cell = ws.cell(row=3, column=1, value="Sistema de Gestión de Parque Automotor — VMT")
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 22

    if LOGO_PATH.exists():
        img = XLImage(str(LOGO_PATH))
        # Logo original ~1986x163 → mantener proporción
        img.width = 520
        img.height = 43
        ws.add_image(img, "A1")

    return 5  # contenido a partir de la fila 5


def _escribir_resumenes(ws, start_row: int, rows: List[Dict[str, Any]], resumenes_sel: List[str],
                        border_thin, resumen_fill, resumen_title_font, resumen_label_font, resumen_value_font) -> int:
    ws.cell(row=start_row, column=1, value="RESUMEN DEL REPORTE").font = resumen_title_font
    resumen_rows = _calcular_resumenes(rows, resumenes_sel)
    for i, (label, value) in enumerate(resumen_rows):
        r = start_row + 1 + i
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_v = ws.cell(row=r, column=2, value=value)
        cell_l.font = resumen_label_font
        cell_v.font = resumen_value_font
        cell_l.fill = resumen_fill
        cell_v.fill = resumen_fill
        cell_l.border = border_thin
        cell_v.border = border_thin
    return start_row + 1 + len(resumen_rows)


@router.get("/opciones")
async def opciones_reporte(_=Depends(get_current_user)):
    """Catálogo de campos y resúmenes disponibles para el constructor de reportes."""
    return {
        "campos": [
            {"key": k, "label": v["label"], "default": v["default"]}
            for k, v in CAMPOS.items()
        ],
        "resumenes": [
            {"key": k, "label": v["label"], "default": v["default"]}
            for k, v in RESUMENES.items()
        ],
    }


@router.get("/buses/excel")
async def exportar_buses_excel(
    empresas: Optional[str] = Query(None, description="IDs de empresa (hex) separados por coma"),
    estado_bus: Optional[str] = None,
    estado_itv: Optional[str] = None,
    id_marca: Optional[int] = None,
    id_tipo_servicio: Optional[int] = None,
    tipo_servicio: Optional[str] = None,
    año_desde: Optional[int] = None,
    año_hasta: Optional[int] = None,
    campos: Optional[str] = Query(
        None,
        description="Claves de campos separadas por coma. Vacío o omitido con solo_resumen=true = sin detalle.",
    ),
    resumenes: Optional[str] = Query(None, description="Claves de resúmenes separadas por coma"),
    solo_resumen: bool = Query(False, description="Si true, no incluye tabla de detalle (solo resúmenes)"),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Generar Excel personalizado según filtros, campos y resúmenes seleccionados."""
    # solo_resumen / campos="" → sin detalle; campos=None sin flag → defaults
    if solo_resumen or campos == "":
        campos_sel: List[str] = []
    elif campos is None:
        campos_sel = [k for k, v in CAMPOS.items() if v["default"]]
    else:
        campos_sel = [c for c in _parse_list(campos) if c in CAMPOS]

    resumenes_sel = [r for r in _parse_list(resumenes) if r in RESUMENES]

    if not campos_sel and not resumenes_sel:
        raise HTTPException(
            status_code=400,
            detail="Seleccioná al menos un campo del reporte o un resumen.",
        )

    empresas_ids = _parse_list(empresas)

    q = (
        select(Bus)
        .options(
            selectinload(Bus.marca),
            selectinload(Bus.tipo_carroceria),
            selectinload(Bus.marca_carroceria),
            selectinload(Bus.tipo_servicio_rel),
            selectinload(Bus.itv_registros),
            selectinload(Bus.asignaciones),
        )
    )

    filters = []
    if estado_bus:
        est = estado_bus.upper()
        if est == "INACTIVO":
            filters.append(Bus.estado_bus.in_(["INACTIVO", "BAJA"]))
        else:
            filters.append(Bus.estado_bus == est)
    if id_marca:
        filters.append(Bus.id_marca == id_marca)
    if id_tipo_servicio:
        filters.append(Bus.id_tipo_servicio == id_tipo_servicio)
    elif tipo_servicio:
        subq_tipo = select(TipoServicio.id_tipo_servicio).where(
            TipoServicio.nombre.ilike(f"%{tipo_servicio}%")
        )
        filters.append(Bus.id_tipo_servicio.in_(subq_tipo))
    if año_desde is not None:
        filters.append(Bus.año >= año_desde)
    if año_hasta is not None:
        filters.append(Bus.año <= año_hasta)
    if empresas_ids:
        subq = (
            select(BusEmpresa.id_bus)
            .where(
                BusEmpresa.fecha_fin_asignacion.is_(None),
                BusEmpresa.id_eot.in_(empresas_ids),
            )
        )
        filters.append(Bus.id_bus.in_(subq))

    if filters:
        q = q.where(and_(*filters))

    q = q.order_by(Bus.numero_orden.asc().nulls_last(), Bus.id_bus.asc())
    buses = (await db.execute(q)).scalars().all()

    # Mapa id_eot → nombre
    eot_ids = set()
    for b in buses:
        for a in b.asignaciones:
            if a.fecha_fin_asignacion is None:
                eot_ids.add(a.id_eot)

    eot_nombres: Dict[str, str] = {}
    if eot_ids:
        eot_rows = (
            await db.execute(
                select(Eot.id_eot_vmt_hex, Eot.eot_nombre).where(Eot.id_eot_vmt_hex.in_(eot_ids))
            )
        ).all()
        eot_nombres = {r[0]: r[1] for r in eot_rows}

    anio_actual = date.today().year
    rows: List[Dict[str, Any]] = []

    for b in buses:
        itvs = sorted(
            b.itv_registros,
            key=lambda x: x.fecha_vencimiento or date.min,
            reverse=True,
        )
        itv = next((i for i in itvs if i.es_vigente), None) or (itvs[0] if itvs else None)
        venc = itv.fecha_vencimiento if itv else None
        estado = calcular_estado_itv(venc)

        if estado_itv and estado != estado_itv.upper():
            continue

        asig = next((a for a in b.asignaciones if a.fecha_fin_asignacion is None), None)
        empresa_nombre = "-"
        if asig:
            empresa_nombre = eot_nombres.get(asig.id_eot) or asig.id_eot or "-"

        rows.append({
            "id_bus": b.id_bus,
            "numero_orden": b.numero_orden or "-",
            "rua": b.rua or "-",
            "numero_chassis": b.numero_chassis or "-",
            "año": b.año if b.año is not None else "-",
            "antiguedad": (anio_actual - b.año) if b.año else "-",
            "marca": b.marca.nombre if b.marca else "-",
            "tipo_carroceria": b.tipo_carroceria.descripcion if b.tipo_carroceria else "-",
            "marca_carroceria": b.marca_carroceria.nombre if b.marca_carroceria else "-",
            "empresa": empresa_nombre,
            "combustible": b.combustible or "-",
            "capacidad_pasajeros": b.capacidad_pasajeros if b.capacidad_pasajeros is not None else "-",
            "cilindrada": b.cilindrada or "-",
            "color": b.color or "-",
            "tipo_servicio": b.tipo_servicio_rel.nombre if b.tipo_servicio_rel else "-",
            "estado_bus": b.estado_bus or "-",
            "tiene_rampa": "Sí" if b.tiene_rampa else "No",
            "itv_vencimiento": venc,
            "itv_estado": estado,
            "fecha_itv": itv.fecha_itv if itv else None,
        })

    # ── Excel ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parque Automotor VMT"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    resumen_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    resumen_title_font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
    resumen_label_font = Font(name="Calibri", size=11, bold=True)
    resumen_value_font = Font(name="Calibri", size=11)

    content_row = _insertar_logo(ws)
    solo_resumen = not campos_sel and bool(resumenes_sel)

    if not solo_resumen:
        headers = [CAMPOS[c]["label"] for c in campos_sel]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=content_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin

        for row_data in rows:
            content_row += 1
            for col_num, key in enumerate(campos_sel, start=1):
                cell = ws.cell(row=content_row, column=col_num, value=_valor_campo(row_data, key))
                cell.border = border_thin
                cell.alignment = align_center

        data_end = content_row
        resumen_start = data_end + 2 if resumenes_sel else data_end
    else:
        resumen_start = content_row

    if resumenes_sel:
        _escribir_resumenes(
            ws,
            resumen_start,
            rows,
            resumenes_sel,
            border_thin,
            resumen_fill,
            resumen_title_font,
            resumen_label_font,
            resumen_value_font,
        )

    max_col = max(ws.max_column or 1, 2)
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, (ws.max_row or 1) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        current = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = max(current, max_len + 3, 14)

    # Ancho mínimo para logo / etiquetas de resumen
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 36)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 18)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        headers={"Content-Disposition": 'attachment; filename="Parque_Automotor_VMT.xlsx"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Planilla ITV (cuadros calculados desde la DB) ───────────

@router.get("/planilla/pestanas")
async def planilla_pestanas(_=Depends(get_current_user)):
    """Lista de pestañas disponibles (equivalente a hojas del Excel ITV)."""
    from app.services.planilla_reportes_db import PESTANAS
    return {"pestanas": PESTANAS, "fuente": "registro_habilitacion + public.eots"}


@router.get("/planilla/reporte/{key}")
async def planilla_reporte_db(
    key: str,
    anio: Optional[int] = Query(None, ge=2000, le=2100),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Devuelve un cuadro de la planilla ITV calculado en vivo desde la base
    (CUADRO DE EDAD, BAJAS, operativos, ITV, faltantes, etc.).
    """
    from app.services.planilla_reportes_db import obtener_reporte
    try:
        return await obtener_reporte(db, key, anio=anio)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar el reporte '{key}': {exc}",
        ) from exc


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join(c if c not in '[]:*?/\\' else " " for c in (name or "Reporte"))
    cleaned = cleaned.strip() or "Reporte"
    return cleaned[:31]


def _safe_filename(name: str) -> str:
    cleaned = "".join(c if c.isalnum() or c in ("-", "_", " ") else "_" for c in (name or "reporte"))
    cleaned = "_".join(cleaned.split()) or "reporte"
    return cleaned[:80]


@router.get("/planilla/reporte/{key}/excel")
async def planilla_reporte_excel(
    key: str,
    anio: Optional[int] = Query(None, ge=2000, le=2100),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Descarga en Excel el cuadro de la pestaña (datos en vivo desde la DB)."""
    from app.services.planilla_reportes_db import obtener_reporte

    try:
        data = await obtener_reporte(db, key, anio=anio)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar el Excel '{key}': {exc}",
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(str(data.get("hoja") or key))

    border_thin = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    filas = list(data.get("filas") or [])
    row_kinds = list(data.get("row_kinds") or [])

    if data.get("layout") == "cuadro_edad_zonal":
        fills = {
            "title": PatternFill(start_color="F4E04D", end_color="F4E04D", fill_type="solid"),
            "dpto": PatternFill(start_color="B7E4A8", end_color="B7E4A8", fill_type="solid"),
            "meta": PatternFill(start_color="9FD4F0", end_color="9FD4F0", fill_type="solid"),
            "banner": PatternFill(start_color="B7E4A8", end_color="B7E4A8", fill_type="solid"),
            "zonal": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "col_header": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
            "subtotal": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "total": PatternFill(start_color="C5D4E8", end_color="C5D4E8", fill_type="solid"),
        }
        for r_idx, row in enumerate(filas, start=1):
            kind = row_kinds[r_idx - 1] if r_idx - 1 < len(row_kinds) else "data"
            cells = list(row) if isinstance(row, (list, tuple)) else [row]
            for c_idx, val in enumerate(cells, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value="" if val is None else val)
                cell.border = border_thin
                cell.alignment = align_center
                if kind == "title":
                    cell.fill = fills["dpto"] if c_idx == len(cells) else fills["title"]
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif kind == "meta" and c_idx == 3:
                    cell.fill = fills["meta"]
                    cell.font = Font(name="Calibri", size=10, bold=True)
                elif kind == "banner":
                    cell.fill = fills["banner"]
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif kind == "zonal":
                    cell.fill = fills["zonal"]
                    cell.font = Font(name="Calibri", size=10, bold=True)
                elif kind == "col_header":
                    cell.fill = fills["col_header"]
                    cell.font = Font(name="Calibri", size=9, bold=True)
                elif kind == "subtotal":
                    cell.fill = fills["subtotal"]
                    cell.font = Font(name="Calibri", size=9, bold=True)
                elif kind == "total":
                    cell.fill = fills["total"]
                    cell.font = Font(name="Calibri", size=10, bold=True)
                elif kind == "leyenda":
                    cell.font = Font(name="Calibri", size=9, bold=True)
                else:
                    cell.font = Font(name="Calibri", size=9)
        # Anchos típicos del cuadro
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 28
        for col_idx in range(4, (ws.max_column or 4) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 6
    else:
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
        note_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        soft_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        headers = list(data.get("headers") or [])
        header_row_idx = None
        for i, row in enumerate(filas):
            if isinstance(row, (list, tuple)) and headers and list(row) == headers:
                header_row_idx = i
                break
            if (
                isinstance(row, (list, tuple))
                and headers
                and len(row) >= 1
                and str(row[0]) == str(headers[0])
                and list(row)[: len(headers)] == headers
            ):
                header_row_idx = i
                break

        for r_idx, row in enumerate(filas, start=1):
            cells = list(row) if isinstance(row, (list, tuple)) else [row]
            for c_idx, val in enumerate(cells, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
                if header_row_idx is not None and r_idx == header_row_idx + 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = soft_border
                elif header_row_idx is not None and r_idx > header_row_idx + 1 and cells:
                    cell.border = soft_border
                    cell.alignment = align_center
                    if isinstance(val, float) and 0 < val < 1:
                        cell.number_format = "0.0%"
                elif r_idx == 1:
                    cell.font = title_font
                elif r_idx == 2:
                    cell.font = note_font

        if header_row_idx is not None:
            ws.freeze_panes = f"A{header_row_idx + 2}"

        max_col = ws.max_column or 1
        for col_idx in range(1, max_col + 1):
            max_len = 0
            for row_idx in range(1, (ws.max_row or 1) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(max_len + 2, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    fname = _safe_filename(str(data.get("hoja") or key))
    fecha = data.get("fecha") or date.today().isoformat()
    anio_rep = data.get("anio")
    if key == "bajas" and anio_rep:
        filename = f"PLANILLA_DE_BAJA_DE_BUSES_{anio_rep}.xlsx"
    else:
        filename = f"Planilla_ITV_{fname}_{fecha}.xlsx"

    return StreamingResponse(
        stream,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Planillas operativas por empresa (una hoja por EOT) ─────

def _escribir_hoja_planilla_parque(ws, data: dict) -> None:
    """Escribe una hoja con layout planilla parque automotor."""
    border_thin = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fills = {
        "title": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
        "col_header": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "footer": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    }
    filas = list(data.get("filas") or [])
    row_kinds = list(data.get("row_kinds") or [])
    date_cols = {6, 7, 8, 12, 13, 14}  # hab, seg pas, seg ter, itv ant, itv, venc (1-based aprox)

    for r_idx, row in enumerate(filas, start=1):
        kind = row_kinds[r_idx - 1] if r_idx - 1 < len(row_kinds) else "data"
        cells = list(row) if isinstance(row, (list, tuple)) else [row]
        for c_idx, val in enumerate(cells, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value="" if val is None else val)
            cell.border = border_thin
            cell.alignment = align_center
            if isinstance(val, date):
                cell.number_format = "DD/MM/YYYY"
            if kind == "title":
                cell.fill = fills["title"]
                cell.font = Font(name="Calibri", size=11, bold=True)
            elif kind == "subtitle":
                cell.font = Font(name="Calibri", size=10, bold=True)
            elif kind == "col_header":
                cell.fill = fills["col_header"]
                cell.font = Font(name="Calibri", size=9, bold=True)
            elif kind == "footer":
                cell.fill = fills["footer"]
                cell.font = Font(name="Calibri", size=9, bold=True)
                if c_idx in (4, 5) and isinstance(val, date):
                    cell.number_format = "DD/MM/YYYY"
            elif kind == "data":
                cell.font = Font(name="Calibri", size=9)
                if c_idx in date_cols and val is not None:
                    cell.number_format = "DD/MM/YYYY"

    header_row_idx = next(
        (i + 1 for i, k in enumerate(row_kinds) if k == "col_header"),
        None,
    )
    if header_row_idx:
        ws.freeze_panes = f"A{header_row_idx + 1}"

    widths = [8, 14, 6, 18, 10, 12, 12, 14, 14, 22, 16, 16, 16, 12, 14, 14, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _workbook_planillas_parque(planillas: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    for data in planillas:
        base = _safe_sheet_name(str(data.get("hoja") or "Empresa"))
        name = base
        n = 2
        while name in used_names:
            suffix = f" {n}"
            name = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
            n += 1
        used_names.add(name)
        ws = wb.create_sheet(title=name)
        _escribir_hoja_planilla_parque(ws, data)
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sin datos")
        ws.cell(row=1, column=1, value="No hay empresas con buses activos para exportar.")
    return wb


@router.get("/planilla/parque/empresas")
async def planilla_parque_empresas(
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Lista empresas con parque activo disponibles para planilla operativa."""
    from app.services.planilla_parque_empresa import listar_empresas_parque

    empresas = await listar_empresas_parque(db)
    return {
        "total": len(empresas),
        "empresas": empresas,
        "fuente": "registro_habilitacion + public.eots",
    }


@router.get("/planilla/parque/excel")
async def planilla_parque_excel(
    empresas: Optional[str] = Query(
        None,
        description="IDs id_eot_vmt_hex separados por coma. Vacío = todas las empresas con parque.",
    ),
    modo: str = Query(
        "empresa",
        description="operativa = subsidio (solo con ITV); empresa = informes (todos + pie ITV)",
        pattern="^(operativa|empresa)$",
    ),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Descarga un libro Excel con una hoja por empresa.
    - operativa: planilla para subsidio (excluye buses sin ITV / vencidas sin fecha)
    - empresa: planilla para informes (incluye vencidas; pie con totales ITV)
    Sin columna POD/RTD.
    """
    from app.services.planilla_parque_empresa import generar_planillas_parque

    ids = _parse_list(empresas) if empresas else None
    planillas = await generar_planillas_parque(db, ids_eot=ids, modo=modo)  # type: ignore[arg-type]
    if not planillas:
        raise HTTPException(
            status_code=404,
            detail="No se encontraron empresas con buses activos para generar planillas.",
        )

    wb = _workbook_planillas_parque(planillas)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    fecha = date.today().isoformat()
    tag = "Operativas" if modo == "operativa" else "Empresa"
    filename = f"Planillas_Parque_{tag}_{fecha}.xlsx"
    return StreamingResponse(
        stream,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/planilla/parque/{id_eot}")
async def planilla_parque_empresa(
    id_eot: str,
    modo: str = Query("empresa", pattern="^(operativa|empresa)$"),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Vista previa JSON de la planilla de una empresa."""
    from app.services.planilla_parque_empresa import generar_planillas_parque

    planillas = await generar_planillas_parque(db, ids_eot=[id_eot], modo=modo)  # type: ignore[arg-type]
    if not planillas:
        raise HTTPException(
            status_code=404,
            detail=f"No hay buses para la empresa '{id_eot}' en modo '{modo}'.",
        )
    return planillas[0]


# ── Planilla Excel (opcional / legado) ──────────────────────

@router.get("/planilla/estado")
async def planilla_estado(_=Depends(get_current_user)):
    """Indica si hay una planilla ITV Excel cargada (legado)."""
    return planilla_viewer.estado_planilla()


@router.get("/planilla/hojas")
async def planilla_hojas(_=Depends(get_current_user)):
    """Lista las hojas de la planilla Excel cargada (legado)."""
    estado = planilla_viewer.estado_planilla()
    if not estado["disponible"]:
        return {"hojas": [], **estado}
    return {"hojas": planilla_viewer.listar_hojas(), **estado}


@router.get("/planilla/hoja")
async def planilla_hoja(
    nombre: str = Query(..., description="Nombre exacto de la hoja Excel"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    _=Depends(get_current_user),
):
    """Devuelve la grilla de una hoja Excel (legado). Preferir /planilla/reporte/{key}."""
    try:
        return planilla_viewer.leer_hoja(nombre, page=page, page_size=page_size)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error al leer la hoja: {exc}") from exc


@router.post("/planilla/cargar")
async def planilla_cargar(
    file: UploadFile = File(...),
    _user=Depends(require_roles(["ADMIN", "SUPERVISOR"])),
):
    """Sube planilla Excel (legado). Los cuadros principales salen de la DB."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Se requiere un archivo .xlsx")
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo demasiado grande (máx. 50 MB)")
    try:
        path = planilla_viewer.guardar_planilla(contents, file.filename)
        hojas = planilla_viewer.listar_hojas(path)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"No se pudo procesar el Excel: {exc}") from exc
    return {
        "status": "ok",
        "filename": file.filename,
        "hojas": hojas,
        "mensaje": f"Planilla '{file.filename}' cargada con {len(hojas)} hojas.",
    }
