import openpyxl
import psycopg2
import os

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
sheet = wb["General"]

headers = [sheet.cell(row=6, column=c).value for c in range(1, 20)]

for r in range(7, sheet.max_row + 1):
    rua = sheet.cell(row=r, column=5).value
    if rua and str(rua).strip().upper() == "OBB184":
        vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
        print("OBB184 in Excel (Row", r, "):")
        for h, v in zip(headers, vals):
            print(f"  {h}: {v}")

conn = psycopg2.connect(
    host=os.getenv("DB_HOST", "168.90.177.232"),
    port=int(os.getenv("DB_PORT", 2024)),
    user=os.getenv("DB_USER", "cid_admin_user"),
    password=os.getenv("DB_PASSWORD", "vmtdmtcidccm"),
    dbname=os.getenv("DB_NAME", "bbdd-monitoreo-cid"),
    sslmode="disable"
)
cur = conn.cursor()

cur.execute("SELECT id_bus, rua, numero_chassis FROM registro_habilitacion.buses WHERE rua = 'OBB184';")
bus = cur.fetchone()
print("\nOBB184 in DB (buses):", bus)

if bus:
    cur.execute("SELECT * FROM registro_habilitacion.itv_bus WHERE id_bus = %s;", (bus[0],))
    print("OBB184 in DB (itv_bus):", cur.fetchall())

cur.close()
conn.close()
