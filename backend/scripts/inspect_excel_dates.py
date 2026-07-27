import openpyxl
from datetime import datetime, date

file_path = "excel.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

header_row_idx = 6
headers = [sheet.cell(row=header_row_idx, column=c).value for c in range(1, 20)]

print("Header Row 6:")
for idx, h in enumerate(headers, 1):
    print(f"  Col {idx}: {h}")

today = date(2026, 7, 24)

# Columns of interest
cols_to_check = {
    8: "HABILITACIÓN",
    9: "SEGURO PASAJEROS",
    10: "SEGURO TERCEROS",
    14: "FECHA VENC. ITV ANTERIOR",
    15: "FECHA DE ITV",
    16: "VENCIMIENTO DE ITV",
    17: "SITUACIÓN DE ITV"
}

stats = {col: {
    "total": 0, "valid": 0, "vencido": 0, "vigente": 0,
    "none": 0, "text_vals": {}, "year_counts": {}
} for col in cols_to_check}

for r in range(header_row_idx + 1, sheet.max_row + 1):
    orden = sheet.cell(row=r, column=1).value
    chasis = sheet.cell(row=r, column=4).value
    rua = sheet.cell(row=r, column=5).value
    
    # If no orden and no chasis and no rua, skip empty trailing rows
    if not orden and not chasis and not rua:
        continue

    for col in cols_to_check:
        val = sheet.cell(row=r, column=col).value
        stats[col]["total"] += 1
        
        parsed_date = None
        if isinstance(val, datetime):
            parsed_date = val.date()
        elif isinstance(val, date):
            parsed_date = val
        elif isinstance(val, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
                try:
                    parsed_date = datetime.strptime(val.strip(), fmt).date()
                    break
                except:
                    pass
        
        if parsed_date:
            stats[col]["valid"] += 1
            yr = parsed_date.year
            stats[col]["year_counts"][yr] = stats[col]["year_counts"].get(yr, 0) + 1
            if parsed_date < today:
                stats[col]["vencido"] += 1
            else:
                stats[col]["vigente"] += 1
        else:
            stats[col]["none"] += 1
            val_str = str(val).strip() if val is not None else "EMPTY"
            stats[col]["text_vals"][val_str] = stats[col]["text_vals"].get(val_str, 0) + 1

print("\n================ DETAILED STATS (As of Today: 2026-07-24) ================")
for col, name in cols_to_check.items():
    s = stats[col]
    print(f"\nCol {col} — [{name}]:")
    print(f"  Total Rows: {s['total']}")
    print(f"  Fechas válidas reconocidas: {s['valid']} (Vencidas: {s['vencido']}, Vigentes: {s['vigente']})")
    print(f"  Sin fecha / texto / vacío: {s['none']}")
    if s["year_counts"]:
        print(f"  Distribución por año de vencimiento/fecha: {dict(sorted(s['year_counts'].items()))}")
    if s["text_vals"]:
        print(f"  Valores no fecha (top 5): {list(s['text_vals'].items())[:5]}")
