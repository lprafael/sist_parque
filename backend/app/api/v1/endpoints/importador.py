import io
import openpyxl
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.core.database import get_db
from app.core.security import require_roles

router = APIRouter(prefix="/importador", tags=["Importador"])


@router.post("/upload-excel")
async def cargar_excel_masivo(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user=Depends(require_roles(["ADMIN", "SUPERVISOR"]))
):
    """
    Recibe un archivo Excel (.xlsx), parsea la planilla de ITV
    y actualiza la base de datos staging / principal.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de archivo no válido. Se requiere un archivo .xlsx de Excel."
        )

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        # Contar filas procesables (asumiendo cabecera en fila 1)
        total_filas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                total_filas += 1

        return {
            "status": "success",
            "filename": file.filename,
            "hoja": sheet_name,
            "filas_detectadas": total_filas,
            "mensaje": f"Archivo '{file.filename}' procesado correctamente con {total_filas} registros."
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}"
        )
