import openpyxl
from datetime import datetime, date, timedelta

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

today = date(2026, 7, 24)
in_30 = today + timedelta(days=30)

vigentes_cnt = 0
por_vencer_cnt = 0
vencidos_cnt = 0
sin_fecha_cnt = 0
total_excel_rows = 0

def parse_date(val):
    if val is None: return None
    if isinstance(val, (date, datetime)): return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        val = val.strip()
        if val in ("00/00/00", "", "N/A", "-", "None"): return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try: return datetime.strptime(val, fmt).date()
            except ValueError: continue
    return None

for r in range(7, ws.max_row + 1):
    rua = ws.cell(row=r, column=5).value
    chassis = ws.cell(row=r, column=4).value
    if not rua and not chassis:
        continue
        
    total_excel_rows += 1
    venc_raw = ws.cell(row=r, column=16).value
    venc = parse_date(venc_raw)

    if not venc:
        sin_fecha_cnt += 1
    elif venc > in_30:
        vigentes_cnt += 1
    elif venc >= today and venc <= in_30:
        por_vencer_cnt += 1
    else: # venc < today
        vencidos_cnt += 1

print("================ ANÁLISIS EXACTO DEL ARCHIVO EXCEL ================")
print(f"Fecha de Referencia Utilizada : {today}")
print(f"Límite Próximos 30 Días       : {in_30}")
print("-------------------------------------------------------------------")
print(f"1. ITV Vigentes (> {in_30})        : {vigentes_cnt}")
print(f"2. ITV Por Vencer (entre {today} y {in_30}): {por_vencer_cnt}")
print(f"3. ITV Vencidas (< {today})         : {vencidos_cnt}")
print(f"4. Sin Fecha / Texto (00/00/00)   : {sin_fecha_cnt}")
print("-------------------------------------------------------------------")
print(f"SUMA TOTAL DE REGISTROS EN EXCEL   : {vigentes_cnt + por_vencer_cnt + vencidos_cnt + sin_fecha_cnt}")
print(f"FILAS LEÍDAS CON RUA/CHASSIS       : {total_excel_rows}")
