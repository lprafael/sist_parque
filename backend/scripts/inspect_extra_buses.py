import openpyxl
import psycopg2
import os

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

excel_ruas = set()
excel_chassis = set()

for r in range(7, 1875):
    rua = ws.cell(row=r, column=5).value
    chassis = ws.cell(row=r, column=4).value
    if rua: excel_ruas.add(str(rua).strip().upper())
    if chassis: excel_chassis.add(str(chassis).strip().upper())

conn = psycopg2.connect(
    host=os.getenv("DB_HOST", "168.90.177.232"),
    port=int(os.getenv("DB_PORT", 2024)),
    user=os.getenv("DB_USER", "cid_admin_user"),
    password=os.getenv("DB_PASSWORD", "vmtdmtcidccm"),
    dbname=os.getenv("DB_NAME", "bbdd-monitoreo-cid"),
    sslmode="disable"
)
cur = conn.cursor()

cur.execute("SELECT id_bus, rua, numero_chassis, fecha_registro FROM registro_habilitacion.buses;")
db_buses = cur.fetchall()

extra_buses = []
for id_b, rua_b, ch_b, f_reg in db_buses:
    rua_s = str(rua_b).strip().upper() if rua_b else ""
    ch_s = str(ch_b).strip().upper() if ch_b else ""
    
    if rua_s not in excel_ruas and ch_s not in excel_chassis:
        extra_buses.append((id_b, rua_b, ch_b, f_reg))

print(f"Total Buses en DB: {len(db_buses)}")
print(f"Total Buses Únicos en Excel: {len(excel_ruas)}")
print(f"Buses Extra en DB que NO pertenecen al Excel oficial 2026: {len(extra_buses)}")

print("\nMuestra de 10 buses extra en la DB (buses obsoletos o de prueba):")
for b in extra_buses[:10]:
    print(" ", b)

cur.close()
conn.close()
