import openpyxl
import psycopg2
import os

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

conn = psycopg2.connect(
    host=os.getenv("DB_HOST", "168.90.177.232"),
    port=int(os.getenv("DB_PORT", 2024)),
    user=os.getenv("DB_USER", "cid_admin_user"),
    password=os.getenv("DB_PASSWORD", "vmtdmtcidccm"),
    dbname=os.getenv("DB_NAME", "bbdd-monitoreo-cid"),
    sslmode="disable"
)
cur = conn.cursor()

cur.execute("SELECT id_bus, rua, numero_chassis FROM registro_habilitacion.buses;")
db_buses = cur.fetchall()

db_rua_map = {}
db_chassis_map = {}

for id_b, rua_b, ch_b in db_buses:
    if rua_b: db_rua_map[str(rua_b).strip().upper()] = id_b
    if ch_b: db_chassis_map[str(ch_b).strip().upper()] = id_b

print(f"Total Buses in DB: {len(db_buses)}")
print(f"Unique DB RUAs: {len(db_rua_map)}")
print(f"Unique DB Chassis: {len(db_chassis_map)}")

matched_rua = 0
matched_chassis = 0
not_matched = 0

sample_unmatched = []

for r in range(7, ws.max_row + 1):
    rua = ws.cell(row=r, column=5).value
    chassis = ws.cell(row=r, column=4).value
    if not rua and not chassis: continue
    
    rua_s = str(rua).strip().upper() if rua else None
    ch_s = str(chassis).strip().upper() if chassis else None
    
    if rua_s and rua_s in db_rua_map:
        matched_rua += 1
    elif ch_s and ch_s in db_chassis_map:
        matched_chassis += 1
    else:
        not_matched += 1
        if len(sample_unmatched) < 5:
            sample_unmatched.append((r, rua_s, ch_s))

print(f"\nExcel Matching Stats against DB buses:")
print(f"  Matched by RUA: {matched_rua}")
print(f"  Matched by Chassis: {matched_chassis}")
print(f"  Not Matched in DB: {not_matched}")
print(f"\nSample Unmatched Excel Rows: {sample_unmatched}")

cur.close()
conn.close()
