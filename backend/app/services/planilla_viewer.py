"""Lectura de hojas de la planilla ITV Excel para visualización en /reportes."""
from __future__ import annotations

import io
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
PLANILLA_PATH = DATA_DIR / "planilla_itv.xlsx"

# Hojas típicas (orden de pestañas). Si faltan en el archivo, se omiten.
HOJAS_PREFERIDAS = [
    "CUADRO DE EDAD",
    "BAJAS",
    "BUSES OPER RESER Y DECLAR",
    "PORCENTAJE INCLUSIVO",
    "GRAFICOS",
    "PORCENTAJE OPERATIVO ITV APROBA",
    "PORCENTAJE OPER RESOL SOBRE DEC",
    "CANTIDAD FALTANTE",
    "PLANILLA DE PARCIALES",
    "BUSES ELECTRICOS",
    "General",
]


def resolve_planilla_path() -> Optional[Path]:
    env = os.getenv("PLANILLA_ITV_PATH")
    if env and Path(env).is_file():
        return Path(env)
    if PLANILLA_PATH.is_file():
        return PLANILLA_PATH
    # Fallbacks útiles en dev
    candidates = [
        Path.cwd() / "ITV - 2026 Base de Datos 29-07-26.xlsx",
        Path.cwd().parent / "ITV - 2026 Base de Datos 29-07-26.xlsx",
        Path(__file__).resolve().parents[3] / "ITV - 2026 Base de Datos 29-07-26.xlsx",
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def guardar_planilla(file_bytes: bytes, filename: str = "planilla_itv.xlsx") -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    dest = PLANILLA_PATH
    dest.write_bytes(file_bytes)
    meta = DATA_DIR / "planilla_itv.meta.txt"
    meta.write_text(f"{filename}\n{datetime.now().isoformat()}", encoding="utf-8")
    return dest


def estado_planilla() -> dict:
    path = resolve_planilla_path()
    meta_name = None
    meta_fecha = None
    meta = DATA_DIR / "planilla_itv.meta.txt"
    if meta.is_file():
        lines = meta.read_text(encoding="utf-8").strip().splitlines()
        if lines:
            meta_name = lines[0]
        if len(lines) > 1:
            meta_fecha = lines[1]
    return {
        "disponible": path is not None,
        "path": str(path) if path else None,
        "filename": meta_name or (path.name if path else None),
        "cargada_en": meta_fecha,
        "tamaño_mb": round(path.stat().st_size / (1024 * 1024), 2) if path else None,
    }


def listar_hojas(path: Optional[Path] = None) -> list[str]:
    path = path or resolve_planilla_path()
    if not path:
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    ordered = [h for h in HOJAS_PREFERIDAS if h in names]
    for n in names:
        if n not in ordered:
            ordered.append(n)
    return ordered


def _cell_value(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, float):
        # porcentajes típicos 0–1
        if 0 < abs(val) <= 1.0001:
            return round(val, 6)
        if val == int(val):
            return int(val)
        return round(val, 4)
    if isinstance(val, str):
        s = val.strip()
        return s if s else None
    return val


def leer_hoja(
    nombre: str,
    *,
    path: Optional[Path] = None,
    page: int = 1,
    page_size: int = 100,
    max_cols: int = 45,
) -> dict:
    path = path or resolve_planilla_path()
    if not path:
        raise FileNotFoundError(
            "No hay planilla ITV cargada. Subí el Excel desde Reportes o configurá PLANILLA_ITV_PATH."
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if nombre not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        raise KeyError(f"Hoja '{nombre}' no encontrada. Disponibles: {', '.join(available)}")

    ws = wb[nombre]
    raw_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_value(c) for c in row[:max_cols]]
        # recortar trailing vacíos
        while cells and cells[-1] is None:
            cells.pop()
        if any(c is not None for c in cells):
            raw_rows.append(cells)

    wb.close()

    total = len(raw_rows)
    page = max(1, page)
    page_size = max(10, min(page_size, 500))
    start = (page - 1) * page_size
    end = start + page_size
    page_rows = raw_rows[start:end]

    # ancho de columnas = máximo de la página (o de todo si pedís mucho)
    width = max((len(r) for r in page_rows), default=0)

    # normalizar filas al mismo ancho
    normalized = [r + [None] * (width - len(r)) for r in page_rows]

    titulo = None
    for r in raw_rows[:8]:
        for c in r:
            if isinstance(c, str) and len(c) > 8 and not re.match(r"^\d+$", c):
                titulo = c
                break
        if titulo:
            break

    return {
        "hoja": nombre,
        "titulo": titulo or nombre,
        "total_filas": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
        "columnas": width,
        "filas": normalized,
        "filename": path.name,
    }


def leer_hoja_from_bytes(
    file_bytes: bytes,
    nombre: str,
    *,
    page: int = 1,
    page_size: int = 100,
) -> dict:
    """Parsea sin persistir (preview puntual)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if nombre not in wb.sheetnames:
        names = wb.sheetnames
        wb.close()
        raise KeyError(f"Hoja '{nombre}' no encontrada. Disponibles: {', '.join(names)}")
    # Guardar temporalmente en memoria vía guardar no; reusar lógica leyendo path no aplica.
    # Inline copy of read for bytes:
    ws = wb[nombre]
    raw_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_value(c) for c in row[:45]]
        while cells and cells[-1] is None:
            cells.pop()
        if any(c is not None for c in cells):
            raw_rows.append(cells)
    names = wb.sheetnames
    wb.close()

    total = len(raw_rows)
    page = max(1, page)
    page_size = max(10, min(page_size, 500))
    start = (page - 1) * page_size
    page_rows = raw_rows[start : start + page_size]
    width = max((len(r) for r in page_rows), default=0)
    normalized = [r + [None] * (width - len(r)) for r in page_rows]
    return {
        "hoja": nombre,
        "titulo": nombre,
        "total_filas": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
        "columnas": width,
        "filas": normalized,
        "hojas_disponibles": names,
    }
