import openpyxl
import psycopg2
import os

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

conn = psycopg2.connect(
    host=os.getenv("DB_HOST", "168.90.177.232"),
    port=int(os.getenv("DB_PORT", 2024)),
    user=os.getenv("DB_USER", "cid_admin_user"),
    password=os.getenv("DB_PASSWORD", "vmtdmtcidccm"),
    dbname=os.getenv("DB_NAME", "bbdd-monitoreo-cid"),
    sslmode="disable"
)
cur = conn.cursor()

cur.execute("SELECT id_bus, rua, numero_chassis FROM registro_habilitacion.buses;")
db_buses = cur.fetchall()

rua_map = {}
chassis_map = {}

for id_b, rua_b, ch_b in db_buses:
    if rua_b: rua_map[str(rua_b).strip().upper()] = id_b
    if ch_b: chassis_map[str(ch_b).strip().upper()] = id_b

matched_buses = set()
excel_ruas = set()

for r in range(7, ws.max_row + 1):
    rua = ws.cell(row=r, column=5).value
    chassis = ws.cell(row=r, column=4).value
    if not rua and not chassis: continue
    
    rua_s = str(rua).strip().upper() if rua else None
    ch_s = str(chassis).strip().upper() if chassis else None
    
    if rua_s: excel_ruas.add(rua_s)
    
    id_bus = None
    if rua_s and rua_s in rua_map: id_bus = rua_map[rua_s]
    elif ch_s and ch_s in chassis_map: id_bus = chassis_map[ch_s]
    
    if id_bus:
        matched_buses.add(id_bus)

print(f"Total Rows in Excel with RUA/chassis: {ws.max_row - 6}")
print(f"Total Unique RUAs in Excel: {len(excel_ruas)}")
print(f"Total Unique Matched id_bus in DB: {len(matched_buses)}")

cur.close()
conn.close()
