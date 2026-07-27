import openpyxl
import psycopg2
import os

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
sheet = wb["General"]

excel_ruas = set()
for r in range(7, sheet.max_row + 1):
    rua = sheet.cell(row=r, column=5).value
    if rua:
        excel_ruas.add(str(rua).strip().upper())

print(f"Total RUAs únicas en Excel: {len(excel_ruas)}")

conn = psycopg2.connect(
    host=os.getenv("DB_HOST", "168.90.177.232"),
    port=int(os.getenv("DB_PORT", 2024)),
    user=os.getenv("DB_USER", "cid_admin_user"),
    password=os.getenv("DB_PASSWORD", "vmtdmtcidccm"),
    dbname=os.getenv("DB_NAME", "bbdd-monitoreo-cid"),
    sslmode="disable"
)
cur = conn.cursor()

cur.execute("SELECT id_bus, rua, numero_chassis FROM registro_habilitacion.buses;")
db_buses = cur.fetchall()
print(f"Total Buses en DB: {len(db_buses)}")

db_ruas = {str(b[1]).strip().upper(): b[0] for b in db_buses if b[1]}
print(f"Total RUAs únicas en DB: {len(db_ruas)}")

coincidentes = excel_ruas.intersection(db_ruas.keys())
print(f"RUAs coincidentes entre Excel y DB: {len(coincidentes)}")

solo_excel = excel_ruas - db_ruas.keys()
print(f"RUAs en Excel que NO están en DB: {len(solo_excel)}")

# Sample matching RUA
sample_rua = list(coincidentes)[0] if coincidentes else None
if sample_rua:
    sample_id = db_ruas[sample_rua]
    cur.execute("SELECT * FROM registro_habilitacion.itv_bus WHERE id_bus = %s;", (sample_id,))
    print(f"\nITVs para bus RUA={sample_rua} (id_bus={sample_id}):", cur.fetchall())

cur.close()
conn.close()
