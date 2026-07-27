import openpyxl
from datetime import datetime, date

wb = openpyxl.load_workbook("excel.xlsx", data_only=True)
ws = wb["General"]

valid = 0
none_cnt = 0
types_cnt = {}

for r in range(7, ws.max_row + 1):
    val = ws.cell(row=r, column=16).value # Col 16 is VENCIMIENTO DE ITV
    t_name = type(val).__name__
    types_cnt[t_name] = types_cnt.get(t_name, 0) + 1
    
    if isinstance(val, (date, datetime)):
        valid += 1
    elif isinstance(val, str):
        val_s = val.strip()
        parsed = None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
            try:
                parsed = datetime.strptime(val_s, fmt).date()
                break
            except:
                pass
        if parsed: valid += 1
        else: none_cnt += 1
    else:
        none_cnt += 1

print("Types seen in Col 16 (VENCIMIENTO DE ITV):", types_cnt)
print(f"Total Rows: {ws.max_row - 6}")
print(f"Valid Parsed Dates: {valid}")
print(f"None / Invalid: {none_cnt}")
